import customtkinter as ctk
import sqlite3
import os
import json
import shutil
from datetime import datetime, date
from tkinter import messagebox, filedialog
import tkinter as tk

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

# ── Tema ──────────────────────────────────────────────────────────────────────
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# ── Paleta ────────────────────────────────────────────────────────────────────
C = {
    "bg":        "#0a1628",
    "panel":     "#0f1e38",
    "card":      "#122040",
    "border":    "#1e3560",
    "accent":    "#2563eb",
    "accent2":   "#7c3aed",
    "text":      "#d0e4ff",
    "muted":     "#5a7db0",
    "alta":      "#ef4444",
    "media":     "#f59e0b",
    "baixa":     "#22c55e",
    "done_fg":   "#3a5a8a",
    "hover":     "#1a2f55",
    "vencida":   "#ef4444",
    "hoje":      "#f59e0b",
}

PRIORITY_COLORS  = {"Alta": C["alta"],  "Media": C["media"], "Baixa": C["baixa"],
                    "Média": C["media"]}
PRIORITY_BG      = {"Alta": "#3b0f0f",  "Media": "#3b2a0a",  "Baixa": "#0a2e14",
                    "Média": "#3b2a0a"}
CATEGORY_COLORS  = {
    "Geral":    "#5a7db0",
    "Trabalho": "#2563eb",
    "Pessoal":  "#7c3aed",
    "Compras":  "#f59e0b",
    "Saude":    "#22c55e",
    "Estudos":  "#06b6d4",
}
CATEGORIES = list(CATEGORY_COLORS.keys())


# ══════════════════════════════════════════════════════════════════════════════
#  BANCO DE DADOS
# ══════════════════════════════════════════════════════════════════════════════
class Database:
    def __init__(self, path="tasks.db"):
        self.path = path
        self._init()

    def _conn(self):
        return sqlite3.connect(self.path)

    def _init(self):
        with self._conn() as c:
            c.execute("""
                CREATE TABLE IF NOT EXISTS tasks (
                    id        INTEGER PRIMARY KEY AUTOINCREMENT,
                    name      TEXT    NOT NULL,
                    date      TEXT    DEFAULT '',
                    priority  TEXT    DEFAULT 'Media',
                    category  TEXT    DEFAULT 'Geral',
                    done      INTEGER DEFAULT 0,
                    created   TEXT    DEFAULT ''
                )
            """)
            cols = [r[1] for r in c.execute("PRAGMA table_info(tasks)")]
            if "category" not in cols:
                c.execute("ALTER TABLE tasks ADD COLUMN category TEXT DEFAULT 'Geral'")

    def add(self, name, date_val, priority, category):
        with self._conn() as c:
            c.execute(
                "INSERT INTO tasks (name,date,priority,category,done,created) VALUES (?,?,?,?,0,?)",
                (name, date_val, priority, category,
                 datetime.now().strftime("%Y-%m-%d %H:%M"))
            )

    def all(self):
        with self._conn() as c:
            return c.execute(
                "SELECT id,name,date,priority,category,done FROM tasks ORDER BY id DESC"
            ).fetchall()

    def toggle(self, tid):
        with self._conn() as c:
            c.execute("UPDATE tasks SET done=1-done WHERE id=?", (tid,))

    def delete(self, tid):
        with self._conn() as c:
            c.execute("DELETE FROM tasks WHERE id=?", (tid,))

    def update(self, tid, name, date_val, priority, category):
        with self._conn() as c:
            c.execute(
                "UPDATE tasks SET name=?,date=?,priority=?,category=? WHERE id=?",
                (name, date_val, priority, category, tid)
            )

    def to_json(self):
        rows = self.all()
        return [
            {"id": r[0], "name": r[1], "date": r[2],
             "priority": r[3], "category": r[4], "done": bool(r[5])}
            for r in rows
        ]


# ══════════════════════════════════════════════════════════════════════════════
#  BACKUP AUTOMATICO
# ══════════════════════════════════════════════════════════════════════════════
def auto_backup(db: Database):
    backup_dir = "backups"
    os.makedirs(backup_dir, exist_ok=True)
    week = datetime.now().strftime("%Y-W%U")
    path = os.path.join(backup_dir, f"backup_{week}.json")
    if not os.path.exists(path):
        data = {"generated": datetime.now().isoformat(), "tasks": db.to_json()}
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS DE DATA
# ══════════════════════════════════════════════════════════════════════════════
def fmt_date(d):
    if not d:
        return "Sem data"
    try:
        return datetime.strptime(d, "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return d

def date_status(d):
    if not d:
        return None
    try:
        t = date.fromisoformat(d)
        if t < date.today():
            return "vencida"
        if t == date.today():
            return "hoje"
        return "normal"
    except Exception:
        return None


# ══════════════════════════════════════════════════════════════════════════════
#  EXPORTAR EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def export_excel(tasks, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tarefas"

    ws.merge_cells("A1:G1")
    ws["A1"] = "TASK MANAGER - Relatorio de Tarefas"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill = PatternFill("solid", fgColor="0D1F3C")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(size=9, color="7A9CC8", name="Calibri")
    ws["A2"].fill = PatternFill("solid", fgColor="091830")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8

    headers = ["#", "Tarefa", "Categoria", "Prazo", "Prioridade", "Status", "Situacao"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True, size=10, color="D0E4FF", name="Calibri")
        cell.fill = PatternFill("solid", fgColor="1E3560")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color="2563EB"))
    ws.row_dimensions[4].height = 24

    pri_colors  = {"Alta": "EF4444", "Media": "F59E0B", "Baixa": "22C55E", "Media": "F59E0B"}
    date_colors = {"vencida": "EF4444", "hoje": "F59E0B", "normal": "22C55E"}

    for i, (tid, name, d, priority, category, done) in enumerate(tasks, 1):
        row = i + 4
        bg  = "0F1E38" if i % 2 == 0 else "0A1628"
        ds  = date_status(d)
        sit = {"vencida": "Vencida!", "hoje": "Vence hoje"}.get(ds or "", "")

        vals = [i, name, category or "Geral", fmt_date(d), priority,
                "Concluida" if done else "Pendente", sit]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(size=10, name="Calibri",
                             color="3A5A8A" if done else "D0E4FF", strike=bool(done))
            cell.alignment = Alignment(
                horizontal="center" if col != 2 else "left", vertical="center")

        ws.cell(row=row, column=5).font = Font(
            bold=True, size=10, name="Calibri",
            color=pri_colors.get(priority, "D0E4FF"))
        ws.cell(row=row, column=6).font = Font(
            bold=True, size=10, name="Calibri",
            color="22C55E" if done else "F59E0B")
        if ds and ds != "normal" and not done:
            ws.cell(row=row, column=7).font = Font(
                bold=True, size=10, name="Calibri",
                color=date_colors.get(ds, "D0E4FF"))
        ws.row_dimensions[row].height = 22

    total      = len(tasks)
    done_count = sum(1 for t in tasks if t[5])
    pending    = total - done_count
    summary_row = len(tasks) + 6
    ws.row_dimensions[summary_row - 1].height = 10

    for offset, (label, val, color) in enumerate([
        ("Total de tarefas", total,      "2563EB"),
        ("Concluidas",       done_count, "22C55E"),
        ("Pendentes",        pending,    "F59E0B"),
    ]):
        r = summary_row + offset
        ws.merge_cells(f"A{r}:E{r}")
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(size=10, name="Calibri", color="7A9CC8")
        ws[f"A{r}"].fill = PatternFill("solid", fgColor="0F1E38")
        ws[f"A{r}"].alignment = Alignment(horizontal="right")
        ws[f"F{r}"] = val
        ws[f"F{r}"].font = Font(bold=True, size=11, name="Calibri", color=color)
        ws[f"F{r}"].fill = PatternFill("solid", fgColor="0F1E38")
        ws[f"F{r}"].alignment = Alignment(horizontal="center")
        ws.merge_cells(f"F{r}:G{r}")
        ws.row_dimensions[r].height = 20

    for col, width in zip(range(1, 8), [6, 36, 14, 16, 14, 14, 14]):
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(path)


# ══════════════════════════════════════════════════════════════════════════════
#  EXPORTAR PDF
# ══════════════════════════════════════════════════════════════════════════════
class PDF(FPDF):
    def header(self):
        self.set_fill_color(13, 31, 60)
        self.rect(0, 0, 210, 32, "F")
        self.set_font("Helvetica", "B", 16)
        self.set_text_color(208, 228, 255)
        self.set_y(8)
        self.cell(0, 10, "Task Manager", align="C", new_x="LMARGIN", new_y="NEXT")
        self.set_font("Helvetica", "", 9)
        self.set_text_color(90, 125, 176)
        self.cell(0, 6,
                  f"Relatorio de Tarefas  -  {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                  align="C", new_x="LMARGIN", new_y="NEXT")
        self.ln(4)

    def footer(self):
        self.set_y(-14)
        self.set_font("Helvetica", "", 8)
        self.set_text_color(58, 90, 138)
        self.cell(0, 10,
                  f"Feito com dedicacao por Murilo Silva  -  Pagina {self.page_no()}",
                  align="C")


def export_pdf(tasks, path):
    pdf = PDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=18)

    total      = len(tasks)
    done_count = sum(1 for t in tasks if t[5])
    pending    = total - done_count

    pdf.set_fill_color(18, 32, 64)
    pdf.set_font("Helvetica", "B", 9)
    for label, val, r, g, b in [
        ("Total",      total,      37,  99, 235),
        ("Concluidas", done_count, 34, 197,  94),
        ("Pendentes",  pending,   245, 158,  11),
    ]:
        pdf.set_text_color(90, 125, 176)
        pdf.cell(30, 9, label, border=1, align="C", fill=True)
        pdf.set_text_color(r, g, b)
        pdf.cell(10, 9, str(val), border=1, align="C", fill=True)
        pdf.cell(4,  9, "",        border=0)
    pdf.ln(14)

    col_widths = [8, 54, 22, 24, 22, 26, 26]
    headers    = ["#", "Tarefa", "Categoria", "Prazo", "Prioridade", "Status", "Situacao"]
    pdf.set_fill_color(30, 53, 96)
    pdf.set_text_color(208, 228, 255)
    pdf.set_font("Helvetica", "B", 8)
    for w, h in zip(col_widths, headers):
        pdf.cell(w, 10, h, border=1, align="C", fill=True)
    pdf.ln()

    pri_colors  = {"Alta": (239,68,68), "Media": (245,158,11), "Baixa": (34,197,94),
                   "Média": (245,158,11)}
    date_colors = {"vencida": (239,68,68), "hoje": (245,158,11)}

    for i, (tid, name, d, priority, category, done) in enumerate(tasks, 1):
        fill = i % 2 == 0
        bg   = (15, 30, 56) if fill else (10, 22, 40)
        ds   = date_status(d)
        sit  = {"vencida": "Vencida!", "hoje": "Vence hoje"}.get(ds or "", "Ok")
        pdf.set_fill_color(*bg)

        task_name = name if len(name) <= 28 else name[:26] + ".."
        cat       = (category or "Geral")[:10]

        pdf.set_text_color(90, 125, 176)
        pdf.set_font("Helvetica", "", 8)
        pdf.cell(col_widths[0], 9, str(i), border=1, align="C", fill=True)

        pdf.set_text_color(*(58,90,138) if done else (208,228,255))
        pdf.set_font("Helvetica", "I" if done else "", 8)
        pdf.cell(col_widths[1], 9, task_name, border=1, fill=True)

        pdf.set_text_color(90, 125, 176)
        pdf.set_font("Helvetica", "", 8)
        pdf.cell(col_widths[2], 9, cat,        border=1, align="C", fill=True)
        pdf.cell(col_widths[3], 9, fmt_date(d), border=1, align="C", fill=True)

        pr, pg, pb = pri_colors.get(priority, (90,125,176))
        pdf.set_text_color(pr, pg, pb)
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(col_widths[4], 9, priority, border=1, align="C", fill=True)

        sc = (34,197,94) if done else (245,158,11)
        pdf.set_text_color(*sc)
        pdf.cell(col_widths[5], 9, "Concluida" if done else "Pendente",
                 border=1, align="C", fill=True)

        if ds in date_colors and not done:
            pdf.set_text_color(*date_colors[ds])
        else:
            pdf.set_text_color(90, 125, 176)
        pdf.set_font("Helvetica", "B" if ds in ("vencida","hoje") else "", 8)
        pdf.cell(col_widths[6], 9, sit, border=1, align="C", fill=True)
        pdf.ln()

    pdf.output(path)


# ══════════════════════════════════════════════════════════════════════════════
#  DIALOG DE EDICAO
# ══════════════════════════════════════════════════════════════════════════════
class TaskDialog(ctk.CTkToplevel):
    def __init__(self, parent, on_save, task_data=None):
        super().__init__(parent)
        self.title("Editar Tarefa" if task_data else "Nova Tarefa")
        self.geometry("440x400")
        self.configure(fg_color=C["panel"])
        self.resizable(False, False)
        self.grab_set()
        self.on_save   = on_save
        self.task_data = task_data

        tid, name, d, priority, category, done = (
            task_data if task_data else (None, "", "", "Media", "Geral", False)
        )

        ctk.CTkLabel(self,
                     text="Editar Tarefa" if task_data else "Nova Tarefa",
                     font=("Helvetica", 15, "bold"),
                     text_color=C["text"]).pack(padx=20, pady=(18,10), anchor="w")

        ctk.CTkLabel(self, text="Nome da tarefa", font=("Helvetica", 11),
                     text_color=C["muted"]).pack(padx=20, pady=(6,0), anchor="w")
        self.name_var = ctk.StringVar(value=name)
        ctk.CTkEntry(self, textvariable=self.name_var,
                     placeholder_text="Ex: Comprar impressora...",
                     fg_color=C["card"], border_color=C["border"],
                     text_color=C["text"], width=400, height=38).pack(padx=20, pady=(4,0))

        ctk.CTkLabel(self, text="Data de prazo (AAAA-MM-DD)", font=("Helvetica", 11),
                     text_color=C["muted"]).pack(padx=20, pady=(10,0), anchor="w")
        self.date_var = ctk.StringVar(value=d)
        ctk.CTkEntry(self, textvariable=self.date_var,
                     placeholder_text="Ex: 2026-06-30",
                     fg_color=C["card"], border_color=C["border"],
                     text_color=C["text"], width=400, height=38).pack(padx=20, pady=(4,0))

        row = ctk.CTkFrame(self, fg_color="transparent")
        row.pack(padx=20, pady=(10,0), fill="x")
        row.grid_columnconfigure((0,1), weight=1)

        ctk.CTkLabel(row, text="Prioridade", font=("Helvetica", 11),
                     text_color=C["muted"]).grid(row=0, column=0, sticky="w")
        ctk.CTkLabel(row, text="Categoria", font=("Helvetica", 11),
                     text_color=C["muted"]).grid(row=0, column=1, sticky="w", padx=(12,0))

        self.pri_var = ctk.StringVar(value=priority)
        ctk.CTkOptionMenu(row, values=["Alta","Media","Baixa"],
                          variable=self.pri_var,
                          fg_color=C["card"], button_color=C["accent"],
                          text_color=C["text"], height=36
                          ).grid(row=1, column=0, sticky="ew", pady=(4,0))

        self.cat_var = ctk.StringVar(value=category or "Geral")
        ctk.CTkOptionMenu(row, values=CATEGORIES,
                          variable=self.cat_var,
                          fg_color=C["card"], button_color=C["accent"],
                          text_color=C["text"], height=36
                          ).grid(row=1, column=1, sticky="ew", padx=(12,0), pady=(4,0))

        ctk.CTkButton(self,
                      text="Salvar" if task_data else "Adicionar",
                      command=self._save,
                      fg_color=C["accent"], hover_color=C["accent2"],
                      text_color="white", height=42, width=400,
                      font=("Helvetica", 13, "bold")
                      ).pack(padx=20, pady=(20,8))

    def _save(self):
        name = self.name_var.get().strip()
        if not name:
            messagebox.showwarning("Aviso", "O nome nao pode estar vazio.")
            return
        d = self.date_var.get().strip()
        if d:
            try:
                datetime.strptime(d, "%Y-%m-%d")
            except ValueError:
                messagebox.showerror("Erro", "Data invalida. Use AAAA-MM-DD.")
                return
        tid = self.task_data[0] if self.task_data else None
        self.on_save(tid, name, d, self.pri_var.get(), self.cat_var.get())
        self.destroy()


# ══════════════════════════════════════════════════════════════════════════════
#  BARRA DE PROGRESSO (Canvas nativo)
# ══════════════════════════════════════════════════════════════════════════════
class ProgressBar(tk.Canvas):
    def __init__(self, master, **kw):
        super().__init__(master, height=8, bg=C["panel"],
                         highlightthickness=0, **kw)
        self._pct = 0.0
        self.bind("<Configure>", lambda e: self._draw())

    def set(self, pct):
        self._pct = max(0.0, min(1.0, pct))
        self._draw()

    def _draw(self):
        w = self.winfo_width()
        if w < 2:
            return
        self.delete("all")
        self._rounded_rect(0, 0, w, 8, 4, fill=C["border"], outline="")
        fill_w = int(w * self._pct)
        if fill_w > 4:
            self._rounded_rect(0, 0, fill_w, 8, 4, fill=C["accent"], outline="")

    def _rounded_rect(self, x1, y1, x2, y2, r, **kw):
        pts = [x1+r,y1, x2-r,y1, x2,y1, x2,y1+r,
               x2,y2-r, x2,y2, x2-r,y2, x1+r,y2,
               x1,y2, x1,y2-r, x1,y1+r, x1,y1]
        return self.create_polygon(pts, smooth=True, **kw)


# ══════════════════════════════════════════════════════════════════════════════
#  CARD DE TAREFA
# ══════════════════════════════════════════════════════════════════════════════
class TaskCard(ctk.CTkFrame):
    def __init__(self, master, task_data, on_toggle, on_delete, on_edit, **kw):
        tid, name, d, priority, category, done = task_data
        ds = date_status(d)

        border_color = C["border"]
        border_width = 1
        if not done:
            if ds == "vencida":
                border_color = C["vencida"]
                border_width = 2
            elif ds == "hoje":
                border_color = C["hoje"]
                border_width = 2

        super().__init__(master, fg_color=C["card"], corner_radius=10,
                         border_width=border_width, border_color=border_color, **kw)

        pri_color = PRIORITY_COLORS.get(priority, C["muted"])
        cat_color = CATEGORY_COLORS.get(category, C["muted"])

        self.check_var = ctk.BooleanVar(value=bool(done))
        ctk.CTkCheckBox(
            self, variable=self.check_var,
            text="", width=20, height=20,
            checkbox_width=20, checkbox_height=20,
            corner_radius=5,
            fg_color=C["accent"], hover_color=C["accent2"],
            border_color=C["border"],
            command=lambda: on_toggle(tid)
        ).grid(row=0, column=0, rowspan=2, padx=(12,8), pady=10)

        disp = (name[:40]+"...") if len(name) > 40 else name
        ctk.CTkLabel(
            self, text=disp,
            font=("Helvetica", 13, "normal"),
            text_color=C["done_fg"] if done else C["text"],
            anchor="w"
        ).grid(row=0, column=1, sticky="w", pady=(8,0))

        meta_parts  = [fmt_date(d)]
        meta_color  = C["muted"]
        if not done:
            if ds == "vencida":
                meta_parts.append("Vencida!")
                meta_color = C["vencida"]
            elif ds == "hoje":
                meta_parts.append("Vence hoje!")
                meta_color = C["hoje"]
        ctk.CTkLabel(
            self, text="  -  ".join(meta_parts),
            font=("Helvetica", 10), text_color=meta_color
        ).grid(row=1, column=1, sticky="w", pady=(0,8))

        ctk.CTkLabel(
            self, text=category or "Geral",
            font=("Helvetica", 9, "bold"),
            text_color=cat_color,
            fg_color=C["panel"],
            corner_radius=8, padx=8, pady=2
        ).grid(row=0, column=2, padx=4, pady=(8,0))

        ctk.CTkLabel(
            self, text=priority,
            font=("Helvetica", 10, "bold"),
            text_color=pri_color,
            fg_color=PRIORITY_BG.get(priority, C["card"]),
            corner_radius=12, padx=10, pady=2
        ).grid(row=1, column=2, padx=4, pady=(0,8))

        ctk.CTkButton(
            self, text="E", width=30, height=30,
            fg_color="transparent", hover_color=C["hover"],
            text_color=C["muted"], font=("Helvetica", 13),
            command=lambda: on_edit(task_data)
        ).grid(row=0, column=3, rowspan=2, padx=2)

        ctk.CTkButton(
            self, text="X", width=30, height=30,
            fg_color="transparent", hover_color="#3b1212",
            text_color=C["muted"], font=("Helvetica", 13, "bold"),
            command=lambda: on_delete(tid)
        ).grid(row=0, column=4, rowspan=2, padx=(2,10))

        self.grid_columnconfigure(1, weight=1)


# ══════════════════════════════════════════════════════════════════════════════
#  APLICATIVO PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════
class TaskManagerApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Task Manager")
        self.geometry("1020x720")
        self.minsize(860, 600)
        self.configure(fg_color=C["bg"])

        self.db        = Database()
        self._sort_key = "id"
        self._sort_asc = False

        auto_backup(self.db)
        self._build_ui()
        self.refresh()
        self.after(800, self._check_deadlines)

    # ── UI ────────────────────────────────────────────────────────────────────
    def _build_ui(self):
        self.grid_columnconfigure(0, weight=0, minsize=320)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self._build_header()
        self._build_left()
        self._build_right()
        self._build_footer()

    def _build_header(self):
        hdr = ctk.CTkFrame(self, fg_color=C["panel"], corner_radius=12,
                           border_width=1, border_color=C["border"])
        hdr.grid(row=0, column=0, columnspan=2, padx=16, pady=(16,8), sticky="ew")
        hdr.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(hdr, text="Tarefas", font=("Helvetica", 26)).grid(
            row=0, column=0, padx=(16,10), pady=10, rowspan=3)

        ctk.CTkLabel(hdr, text="Task Manager",
                     font=("Helvetica", 20, "bold"),
                     text_color=C["text"]).grid(row=0, column=1, sticky="w", pady=(10,0))
        ctk.CTkLabel(hdr, text="Organize suas tarefas com eficiencia",
                     font=("Helvetica", 11),
                     text_color=C["muted"]).grid(row=1, column=1, sticky="w")

        prog_frame = ctk.CTkFrame(hdr, fg_color="transparent")
        prog_frame.grid(row=2, column=1, sticky="ew", pady=(4,10), padx=(0,16))
        prog_frame.grid_columnconfigure(0, weight=1)

        self.progress_bar = ProgressBar(prog_frame)
        self.progress_bar.grid(row=0, column=0, sticky="ew", pady=(0,2))
        self.lbl_progress = ctk.CTkLabel(prog_frame, text="0% concluido",
                                         font=("Helvetica", 9),
                                         text_color=C["muted"])
        self.lbl_progress.grid(row=1, column=0, sticky="w")

        stats = ctk.CTkFrame(hdr, fg_color="transparent")
        stats.grid(row=0, column=2, rowspan=3, padx=16, pady=10)
        self.lbl_pending = self._stat_card(stats, "Pendentes", "0", 0)
        self.lbl_done    = self._stat_card(stats, "Concluidas", "0", 1)
        self.lbl_vencida = self._stat_card(stats, "Vencidas",  "0", 2, warn=True)

    def _stat_card(self, parent, label, num, col, warn=False):
        f = ctk.CTkFrame(parent, fg_color=C["card"], corner_radius=10,
                         border_width=1, border_color=C["border"])
        f.grid(row=0, column=col, padx=5)
        inner = ctk.CTkFrame(f, fg_color="transparent")
        inner.pack(padx=14, pady=8)
        color = C["alta"] if warn else C["text"]
        lbl = ctk.CTkLabel(inner, text=num,
                           font=("Helvetica", 20, "bold"), text_color=color)
        lbl.pack()
        ctk.CTkLabel(inner, text=label, font=("Helvetica", 10),
                     text_color=C["muted"]).pack()
        return lbl

    def _build_left(self):
        left = ctk.CTkFrame(self, fg_color="transparent")
        left.grid(row=1, column=0, padx=(16,8), pady=0, sticky="nsew")

        # Formulario
        form = ctk.CTkFrame(left, fg_color=C["panel"], corner_radius=12,
                            border_width=1, border_color=C["border"])
        form.grid(row=0, column=0, sticky="ew", pady=(0,8))
        form.grid_columnconfigure(0, weight=1)

        self._sec(form, "NOVA TAREFA", 0)

        self.entry_name = ctk.CTkEntry(
            form, placeholder_text="Digite uma nova tarefa...",
            fg_color=C["card"], border_color=C["border"],
            text_color=C["text"], placeholder_text_color=C["muted"], height=38)
        self.entry_name.grid(row=1, column=0, padx=14, pady=(0,8), sticky="ew")
        self.entry_name.bind("<Return>", lambda e: self.add_task())

        self.entry_date = ctk.CTkEntry(
            form, placeholder_text="Data (AAAA-MM-DD)",
            fg_color=C["card"], border_color=C["border"],
            text_color=C["text"], placeholder_text_color=C["muted"], height=38)
        self.entry_date.grid(row=2, column=0, padx=14, pady=(0,8), sticky="ew")

        duo = ctk.CTkFrame(form, fg_color="transparent")
        duo.grid(row=3, column=0, padx=14, pady=(0,8), sticky="ew")
        duo.grid_columnconfigure((0,1), weight=1)

        self.pri_var = ctk.StringVar(value="Media")
        ctk.CTkOptionMenu(duo, values=["Alta","Media","Baixa"],
                          variable=self.pri_var,
                          fg_color=C["card"], button_color=C["accent"],
                          button_hover_color=C["accent2"],
                          text_color=C["text"], height=38
                          ).grid(row=0, column=0, sticky="ew", padx=(0,6))

        self.cat_var = ctk.StringVar(value="Geral")
        ctk.CTkOptionMenu(duo, values=CATEGORIES,
                          variable=self.cat_var,
                          fg_color=C["card"], button_color=C["accent"],
                          button_hover_color=C["accent2"],
                          text_color=C["text"], height=38
                          ).grid(row=0, column=1, sticky="ew")

        ctk.CTkButton(
            form, text="Adicionar Tarefa",
            command=self.add_task,
            fg_color=C["accent"], hover_color=C["accent2"],
            text_color="white", height=42,
            font=("Helvetica", 13, "bold")
        ).grid(row=4, column=0, padx=14, pady=(0,14), sticky="ew")

        # Acoes
        act = ctk.CTkFrame(left, fg_color=C["panel"], corner_radius=12,
                           border_width=1, border_color=C["border"])
        act.grid(row=1, column=0, sticky="ew", pady=(0,8))
        act.grid_columnconfigure((0,1), weight=1)

        self._sec(act, "ACOES", 0, span=2)

        ctk.CTkButton(act, text="Exportar Excel",
                      command=self.export_excel,
                      fg_color="#14532d", hover_color="#166534",
                      text_color="#4ade80", height=38,
                      font=("Helvetica", 12, "bold")
                      ).grid(row=1, column=0, padx=(14,6), pady=(0,8), sticky="ew")

        ctk.CTkButton(act, text="Exportar PDF",
                      command=self.export_pdf,
                      fg_color="#7f1d1d", hover_color="#991b1b",
                      text_color="#f87171", height=38,
                      font=("Helvetica", 12, "bold")
                      ).grid(row=1, column=1, padx=(6,14), pady=(0,8), sticky="ew")

        ctk.CTkButton(act, text="Backup JSON",
                      command=self.manual_backup,
                      fg_color="#1e3560", hover_color="#2a4a80",
                      text_color=C["muted"], height=34,
                      font=("Helvetica", 11)
                      ).grid(row=2, column=0, columnspan=2,
                             padx=14, pady=(0,14), sticky="ew")

        # Ordenacao
        sort_f = ctk.CTkFrame(left, fg_color=C["panel"], corner_radius=12,
                              border_width=1, border_color=C["border"])
        sort_f.grid(row=2, column=0, sticky="ew", pady=(0,8))
        sort_f.grid_columnconfigure((0,1,2), weight=1)

        self._sec(sort_f, "ORDENAR POR", 0, span=3)
        for col, (label, key) in enumerate([("Nome","name"),
                                             ("Data","date"),
                                             ("Prioridade","priority")]):
            ctk.CTkButton(sort_f, text=label,
                          command=lambda k=key: self.sort_by(k),
                          fg_color=C["card"], hover_color=C["hover"],
                          text_color=C["muted"], height=32,
                          font=("Helvetica", 11)
                          ).grid(row=1, column=col,
                                 padx=(14 if col==0 else 4, 4 if col<2 else 14),
                                 pady=(0,14), sticky="ew")

    def _build_right(self):
        right = ctk.CTkFrame(self, fg_color=C["panel"], corner_radius=12,
                             border_width=1, border_color=C["border"])
        right.grid(row=1, column=1, padx=(0,16), pady=0, sticky="nsew")
        right.grid_columnconfigure(0, weight=1)
        right.grid_rowconfigure(3, weight=1)

        top = ctk.CTkFrame(right, fg_color="transparent")
        top.grid(row=0, column=0, padx=14, pady=(14,4), sticky="ew")
        top.grid_columnconfigure(0, weight=1)

        self.search_var = ctk.StringVar()
        self.search_var.trace_add("write", lambda *_: self.refresh())
        ctk.CTkEntry(top, textvariable=self.search_var,
                     placeholder_text="Buscar tarefa...",
                     fg_color=C["card"], border_color=C["border"],
                     text_color=C["text"], placeholder_text_color=C["muted"],
                     height=38
                     ).grid(row=0, column=0, padx=(0,8), sticky="ew")

        self.filter_var = ctk.StringVar(value="Todas")
        ctk.CTkOptionMenu(top,
                          values=["Todas","Pendentes","Concluidas",
                                  "Alta","Media","Baixa","Vencidas","Hoje"],
                          variable=self.filter_var,
                          command=lambda _: self.refresh(),
                          fg_color=C["card"], button_color=C["accent"],
                          button_hover_color=C["accent2"],
                          text_color=C["text"], dropdown_fg_color=C["card"],
                          width=150, height=38
                          ).grid(row=0, column=1)

        cat_row = ctk.CTkFrame(right, fg_color="transparent")
        cat_row.grid(row=1, column=0, padx=14, pady=(4,8), sticky="ew")
        ctk.CTkLabel(cat_row, text="Categoria:", font=("Helvetica", 11),
                     text_color=C["muted"]).pack(side="left", padx=(0,8))
        self.cat_filter_var = ctk.StringVar(value="Todas")
        ctk.CTkOptionMenu(cat_row,
                          values=["Todas"] + CATEGORIES,
                          variable=self.cat_filter_var,
                          command=lambda _: self.refresh(),
                          fg_color=C["card"], button_color=C["accent"],
                          button_hover_color=C["accent2"],
                          text_color=C["text"], dropdown_fg_color=C["card"],
                          width=160, height=34
                          ).pack(side="left")

        ctk.CTkLabel(right, text="SUAS TAREFAS",
                     font=("Helvetica", 10, "bold"),
                     text_color=C["muted"]).grid(
            row=2, column=0, padx=14, pady=(0,4), sticky="w")

        self.scroll = ctk.CTkScrollableFrame(
            right, fg_color="transparent",
            scrollbar_button_color=C["border"])
        self.scroll.grid(row=3, column=0, padx=8, pady=(0,12), sticky="nsew")
        self.scroll.grid_columnconfigure(0, weight=1)

    def _build_footer(self):
        foot = ctk.CTkFrame(self, fg_color=C["panel"], corner_radius=10,
                            border_width=1, border_color=C["border"])
        foot.grid(row=2, column=0, columnspan=2, padx=16, pady=(4,14), sticky="ew")
        foot.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(foot,
                     text="Foco, disciplina e organizacao transformam planos em conquistas.",
                     font=("Helvetica", 10), text_color=C["muted"]).grid(
            row=0, column=0, padx=14, pady=8)
        ctk.CTkLabel(foot,
                     text="Feito com dedicacao por Murilo Silva",
                     font=("Helvetica", 10), text_color="#4ade80").grid(
            row=0, column=2, padx=14, pady=8)

    def _sec(self, parent, text, row, span=1):
        ctk.CTkLabel(parent, text=text,
                     font=("Helvetica", 10, "bold"),
                     text_color=C["muted"]).grid(
            row=row, column=0, columnspan=span,
            padx=14, pady=(14,8), sticky="w")

    # ── Logica ────────────────────────────────────────────────────────────────
    def refresh(self):
        for w in self.scroll.winfo_children():
            w.destroy()

        tasks = self.db.all()
        q  = self.search_var.get().lower()
        f  = self.filter_var.get()
        cf = self.cat_filter_var.get()

        filtered = []
        for t in tasks:
            tid, name, d, priority, category, done = t
            if q and q not in name.lower():
                continue
            if f == "Pendentes"  and done:                               continue
            if f == "Concluidas" and not done:                           continue
            if f in ("Alta","Media","Baixa") and priority != f:          continue
            if f == "Vencidas" and (done or date_status(d) != "vencida"): continue
            if f == "Hoje"     and (done or date_status(d) != "hoje"):    continue
            if cf != "Todas" and (category or "Geral") != cf:            continue
            filtered.append(t)

        key_map = {
            "name":     lambda t: t[1].lower(),
            "date":     lambda t: t[2] or "9999",
            "priority": lambda t: ["Alta","Media","Baixa"].index(t[3])
                                   if t[3] in ["Alta","Media","Baixa"] else 9,
            "id":       lambda t: t[0],
        }
        if self._sort_key in key_map:
            filtered.sort(key=key_map[self._sort_key], reverse=not self._sort_asc)

        if not filtered:
            ctk.CTkLabel(self.scroll, text="Nenhuma tarefa encontrada.",
                         font=("Helvetica", 12), text_color=C["muted"]).pack(pady=30)
        else:
            for t in filtered:
                TaskCard(self.scroll, t,
                         on_toggle=self.toggle_task,
                         on_delete=self.delete_task,
                         on_edit=self.edit_task
                         ).pack(fill="x", pady=4, padx=4)

        all_tasks = self.db.all()
        total    = len(all_tasks)
        done_c   = sum(1 for t in all_tasks if t[5])
        pending  = total - done_c
        vencidas = sum(1 for t in all_tasks
                       if not t[5] and date_status(t[2]) == "vencida")

        self.lbl_pending.configure(text=str(pending))
        self.lbl_done.configure(text=str(done_c))
        self.lbl_vencida.configure(text=str(vencidas))

        pct = done_c / total if total else 0
        self.progress_bar.set(pct)
        self.lbl_progress.configure(
            text=f"{int(pct*100)}% concluido  ({done_c}/{total} tarefas)")

    def sort_by(self, key):
        if self._sort_key == key:
            self._sort_asc = not self._sort_asc
        else:
            self._sort_key = key
            self._sort_asc = True
        self.refresh()

    def add_task(self):
        name = self.entry_name.get().strip()
        if not name:
            messagebox.showwarning("Aviso", "Digite o nome da tarefa.")
            return
        d = self.entry_date.get().strip()
        if d:
            try:
                datetime.strptime(d, "%Y-%m-%d")
            except ValueError:
                messagebox.showerror("Erro", "Data invalida. Use AAAA-MM-DD.")
                return
        self.db.add(name, d, self.pri_var.get(), self.cat_var.get())
        self.entry_name.delete(0, "end")
        self.entry_date.delete(0, "end")
        self.pri_var.set("Media")
        self.cat_var.set("Geral")
        self.refresh()

    def toggle_task(self, tid):
        self.db.toggle(tid)
        self.refresh()

    def delete_task(self, tid):
        if messagebox.askyesno("Confirmar", "Deseja excluir esta tarefa?"):
            self.db.delete(tid)
            self.refresh()

    def edit_task(self, task_data):
        def save(tid, name, d, priority, category):
            self.db.update(tid, name, d, priority, category)
            self.refresh()
        TaskDialog(self, on_save=save, task_data=task_data)

    def _check_deadlines(self):
        tasks    = self.db.all()
        vencidas = [t[1] for t in tasks if not t[5] and date_status(t[2]) == "vencida"]
        hoje     = [t[1] for t in tasks if not t[5] and date_status(t[2]) == "hoje"]
        msgs = []
        if vencidas:
            nomes = "\n".join(f"  - {n}" for n in vencidas[:5])
            msgs.append(f"{len(vencidas)} tarefa(s) VENCIDA(S):\n{nomes}")
        if hoje:
            nomes = "\n".join(f"  - {n}" for n in hoje[:5])
            msgs.append(f"{len(hoje)} tarefa(s) vencem HOJE:\n{nomes}")
        if msgs:
            messagebox.showwarning("Atencao — Prazos", "\n\n".join(msgs))

    def manual_backup(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON", "*.json")],
            initialfile=f"backup_{datetime.now().strftime('%Y%m%d_%H%M')}.json",
            title="Salvar backup"
        )
        if not path:
            return
        data = {"generated": datetime.now().isoformat(), "tasks": self.db.to_json()}
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        messagebox.showinfo("Backup", f"Backup salvo em:\n{path}")

    def export_excel(self):
        tasks = self.db.all()
        if not tasks:
            messagebox.showinfo("Aviso", "Nao ha tarefas para exportar.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="tarefas.xlsx"
        )
        if not path:
            return
        try:
            export_excel(tasks, path)
            if messagebox.askyesno("Sucesso", f"Excel exportado!\n\n{path}\n\nAbrir agora?"):
                os.startfile(path) if os.name == "nt" else os.system(f'open "{path}"')
        except Exception as e:
            messagebox.showerror("Erro", str(e))

    def export_pdf(self):
        tasks = self.db.all()
        if not tasks:
            messagebox.showinfo("Aviso", "Nao ha tarefas para exportar.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            initialfile="tarefas.pdf"
        )
        if not path:
            return
        try:
            export_pdf(tasks, path)
            if messagebox.askyesno("Sucesso", f"PDF exportado!\n\n{path}\n\nAbrir agora?"):
                os.startfile(path) if os.name == "nt" else os.system(f'open "{path}"')
        except Exception as e:
            messagebox.showerror("Erro", str(e))


# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    app = TaskManagerApp()
    app.mainloop()
